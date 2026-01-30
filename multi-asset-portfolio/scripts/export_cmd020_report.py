#!/usr/bin/env python3
"""
cmd_020 レポートをExcel形式で出力
"""
import json
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference, BarChart

# パス設定
RESULTS_DIR = Path(__file__).parent.parent / "results"
OUTPUT_PATH = RESULTS_DIR / "cmd020_report.xlsx"


def load_data():
    """結果データを読み込み"""
    data = {}

    # 月次バックテスト
    monthly_path = RESULTS_DIR / "backtest_full_monthly.json"
    if monthly_path.exists():
        with open(monthly_path) as f:
            data['monthly'] = json.load(f)

    # 週次バックテスト
    weekly_path = RESULTS_DIR / "backtest_full_weekly.json"
    if weekly_path.exists():
        with open(weekly_path) as f:
            data['weekly'] = json.load(f)

    # 品質フィルターレポート
    quality_path = RESULTS_DIR / "quality_filter_report_full.json"
    if quality_path.exists():
        with open(quality_path) as f:
            data['quality'] = json.load(f)

    # 改善提案（YAML）
    import yaml
    improvements_path = RESULTS_DIR / "full_portfolio_improvements.yaml"
    if improvements_path.exists():
        with open(improvements_path) as f:
            data['improvements'] = yaml.safe_load(f)

    return data


def create_styles():
    """スタイル定義"""
    return {
        'header': Font(bold=True, size=12, color="FFFFFF"),
        'header_fill': PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        'title': Font(bold=True, size=14),
        'number': Alignment(horizontal='right'),
        'center': Alignment(horizontal='center'),
        'good': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'bad': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        'neutral': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def create_summary_sheet(wb, data, styles):
    """サマリーシート作成"""
    ws = wb.active
    ws.title = "サマリー"

    # タイトル
    ws['A1'] = "cmd_020 包括的バックテストレポート"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"生成日時: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A3'] = "期間: 2010-01-01 ～ 2024-12-31（約15年）"

    # 主要指標テーブル
    ws['A5'] = "主要指標一覧"
    ws['A5'].font = styles['title']

    headers = ['指標', '月次BT', '週次BT', 'SPY', 'QQQ', '60/40', 'AGG']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['border']

    # データ行
    metrics = [
        ('年率リターン', '14.28%', '6.06%', '13.69%', '18.49%', '10.06%', '2.31%'),
        ('Sharpe Ratio', '0.913', '0.261', '0.803', '0.905', '0.998', '0.486'),
        ('最大DD', '-31.49%', '-33.84%', '-33.72%', '-35.12%', '-27.24%', '-18.43%'),
        ('ボラティリティ', '13.46%', '15.56%', '17.05%', '20.44%', '10.09%', '4.75%'),
        ('Calmar Ratio', '0.453', '0.179', '0.406', '0.527', '0.369', '0.125'),
        ('累計リターン', '694%', '140%', '584%', '1172%', '320%', '41%'),
    ]

    for row_idx, row_data in enumerate(metrics, 7):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            if col_idx > 1:
                cell.alignment = styles['number']

    # 目標達成状況
    ws['A15'] = "目標達成状況"
    ws['A15'].font = styles['title']

    goal_headers = ['目標', '基準', '月次BT', '週次BT', '判定']
    for col, header in enumerate(goal_headers, 1):
        cell = ws.cell(row=16, column=col, value=header)
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['border']

    goals = [
        ('SPY超過リターン', '>13.69%', '14.28%', '6.06%', '月次のみ達成'),
        ('Sharpe > 1.0', '>1.0', '0.913', '0.261', '未達'),
        ('MDD < 25%', '<25%', '-31.49%', '-33.84%', '未達'),
        ('60/40超過Sharpe', '>0.998', '0.913', '0.261', '未達'),
    ]

    for row_idx, row_data in enumerate(goals, 17):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            # 判定列の色分け
            if col_idx == 5:
                if '達成' in value and '未達' not in value:
                    cell.fill = styles['good']
                elif '未達' in value:
                    cell.fill = styles['bad']
                else:
                    cell.fill = styles['neutral']

    # 列幅調整
    ws.column_dimensions['A'].width = 20
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_yearly_sheet(wb, data, styles):
    """年別パフォーマンスシート"""
    ws = wb.create_sheet("年別パフォーマンス")

    ws['A1'] = "年別パフォーマンス比較"
    ws['A1'].font = styles['title']

    headers = ['年', '週次BT', 'SPY', 'QQQ', '60/40', '勝敗']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['border']

    yearly_data = [
        (2010, 8.33, 13.14, 18.41, 12.71, '負'),
        (2011, -11.31, 1.90, 3.48, 15.90, '負'),
        (2012, 15.92, 15.99, 18.11, 11.14, '引分'),
        (2013, 47.40, 32.31, 36.63, 12.17, '勝'),
        (2014, 15.81, 13.46, 19.18, 19.31, '引分'),
        (2015, 10.68, 1.23, 9.44, 0.81, '勝'),
        (2016, -14.01, 12.00, 7.10, 8.13, '負'),
        (2017, 21.50, 21.71, 32.66, 16.79, '引分'),
        (2018, -3.48, -4.57, -0.13, -2.86, '引分'),
        (2019, 30.98, 31.22, 38.96, 24.74, '引分'),
        (2020, -17.46, 18.33, 48.41, 21.54, '負'),
        (2021, 2.06, 28.73, 27.42, 14.76, '負'),
        (2022, -11.03, -18.18, -32.58, -22.83, '勝'),
        (2023, 11.25, 26.18, 54.86, 16.86, '負'),
        (2024, 3.44, 25.34, 26.65, 11.42, '負'),
    ]

    for row_idx, row_data in enumerate(yearly_data, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_idx == 1:
                cell.value = value
            elif col_idx < 6:
                cell.value = f"{value}%"
                cell.alignment = styles['number']
            else:
                cell.value = value
                if value == '勝':
                    cell.fill = styles['good']
                elif value == '負':
                    cell.fill = styles['bad']
            cell.border = styles['border']

    # 勝率
    ws['A20'] = "勝率: 3勝 / 15年 = 20%"
    ws['A20'].font = Font(bold=True)

    # 列幅調整
    ws.column_dimensions['A'].width = 8
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_improvements_sheet(wb, data, styles):
    """改善提案シート"""
    ws = wb.create_sheet("改善提案")

    ws['A1'] = "改善提案一覧（優先度順）"
    ws['A1'].font = styles['title']

    headers = ['ID', '提案', '優先度', 'カテゴリ', '期待リターン', '期待Sharpe', '期待MDD', '工数(日)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.alignment = styles['center']
        cell.border = styles['border']

    improvements = [
        ('IMP-001', 'ユニバースサイズ最適化', 'Critical', 'ユニバース', '+3~5%', '+0.2~0.3', '-3~5%', 2),
        ('IMP-002', 'ドローダウン制御アルゴリズム', 'Critical', 'リスク管理', '-1~2%', '+0.1~0.2', '-8~12%', 5),
        ('IMP-003', 'VIXベース動的キャッシュ配分', 'High', 'リスク管理', '-0.5~1%', '+0.15~0.25', '-5~8%', 2),
        ('IMP-004', 'レジーム検出改善', 'High', '戦略強化', '+2~4%', '+0.1~0.2', '-2~4%', 7),
        ('IMP-005', 'リバランス頻度最適化', 'High', '執行最適化', '+1~2%', '+0.1', '中立', 3),
        ('IMP-006', 'ミーンリバージョン戦略追加', 'Medium', '戦略強化', '+1~3%', '+0.1~0.2', '-2~3%', 5),
        ('IMP-007', 'セクターローテーション最適化', 'Medium', 'セクター', '+0.5~1.5%', '+0.05~0.1', '-1~2%', 4),
        ('IMP-008', '地域分散最適化', 'Medium', '地域', '+0.5~1%', '+0.05~0.1', '-1~2%', 6),
        ('IMP-009', 'クオリティファクター追加', 'Low', 'ファクター', '+0.3~0.8%', '+0.03~0.05', '-0.5~1%', 4),
        ('IMP-010', '取引執行最適化', 'Low', '執行最適化', '+0.1~0.3%', '+0.01~0.02', '中立', 7),
    ]

    priority_colors = {
        'Critical': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
        'High': PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
        'Medium': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
        'Low': PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid"),
    }

    for row_idx, row_data in enumerate(improvements, 4):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']
            if col_idx == 3:  # 優先度列
                cell.fill = priority_colors.get(value, styles['neutral'])
            if col_idx == 8:  # 工数
                cell.alignment = styles['center']

    # ロードマップ
    ws['A16'] = "実装ロードマップ"
    ws['A16'].font = styles['title']

    roadmap_headers = ['フェーズ', '期間', '対象', '期待改善']
    for col, header in enumerate(roadmap_headers, 1):
        cell = ws.cell(row=17, column=col, value=header)
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.border = styles['border']

    roadmap = [
        ('Phase 1', '1-2週間', 'IMP-001, 003, 005', 'Sharpe +0.3~0.5, MDD -6~10%'),
        ('Phase 2', '2-4週間', 'IMP-002, 006, 007', 'Sharpe +0.2~0.4（累積）'),
        ('Phase 3', '1-2ヶ月', 'IMP-004, 008, 009, 010', 'Sharpe +0.1~0.2（累積）'),
    ]

    for row_idx, row_data in enumerate(roadmap, 18):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']

    # 列幅調整
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    for col in range(5, 9):
        ws.column_dimensions[get_column_letter(col)].width = 12


def create_recommendations_sheet(wb, data, styles):
    """推奨設定シート"""
    ws = wb.create_sheet("推奨設定")

    ws['A1'] = "プロダクション推奨設定"
    ws['A1'].font = styles['title']

    ws['A3'] = "基本設定"
    ws['A3'].font = Font(bold=True)

    settings = [
        ('リバランス頻度', '月次（毎月第1営業日）'),
        ('ユニバースサイズ', '100-150銘柄（流動性上位）'),
        ('戦略', 'モメンタム + リスクパリティ'),
        ('初期資本', '$100,000'),
        ('取引コスト', '10bps（片道）'),
    ]

    for row_idx, (key, value) in enumerate(settings, 4):
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    ws['A10'] = "リスク管理設定"
    ws['A10'].font = Font(bold=True)

    risk_settings = [
        ('VIXキャッシュ配分', 'VIX>30で60%キャッシュ'),
        ('ドローダウン制御', '-20%超過で80%削減'),
        ('最大ポジション', '1銘柄あたり5%'),
        ('セクター上限', '各セクター20%'),
    ]

    for row_idx, (key, value) in enumerate(risk_settings, 11):
        ws.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)

    # 期待効果
    ws['A17'] = "改善後の期待パフォーマンス"
    ws['A17'].font = styles['title']

    expected_headers = ['指標', '現状', '改善後目標']
    for col, header in enumerate(expected_headers, 1):
        cell = ws.cell(row=18, column=col, value=header)
        cell.font = styles['header']
        cell.fill = styles['header_fill']
        cell.border = styles['border']

    expected = [
        ('年率リターン', '14.28%', '18~22%'),
        ('Sharpe Ratio', '0.913', '1.1~1.3'),
        ('最大DD', '-31.49%', '-18~22%'),
    ]

    for row_idx, row_data in enumerate(expected, 19):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = styles['border']

    # 列幅調整
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15


def main():
    """メイン処理"""
    print("cmd_020 レポートをExcel形式で出力中...")

    # データ読み込み
    data = load_data()

    # ワークブック作成
    wb = Workbook()
    styles = create_styles()

    # 各シート作成
    create_summary_sheet(wb, data, styles)
    create_yearly_sheet(wb, data, styles)
    create_improvements_sheet(wb, data, styles)
    create_recommendations_sheet(wb, data, styles)

    # 保存
    wb.save(OUTPUT_PATH)
    print(f"✅ レポート出力完了: {OUTPUT_PATH}")

    return OUTPUT_PATH


if __name__ == "__main__":
    main()
